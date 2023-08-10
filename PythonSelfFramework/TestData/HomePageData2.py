import openpyxl

class HomePageData2:

    @staticmethod
    def getAllTestData2():
        book = openpyxl.load_workbook(
            "C:\\Users\\Main-Snow\\PycharmProjects\\PythonSelfFramework\\TestData\\PythonDemo.xlsx")
        sheet = book.active

        testcase_column = 1  # Assuming test cases are in the first column

        # Initialize a list to store dictionaries for each testcase
        testcase_dicts = []

        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=testcase_column).value:
                testcase_dict = {}
                for j in range(2, sheet.max_column + 1):
                    key = sheet.cell(row=1, column=j).value
                    value = sheet.cell(row=i, column=j).value
                    testcase_dict[key] = value
                testcase_dicts.append(testcase_dict)

        return testcase_dicts
