import openpyxl


class HomePageData:

        test_HomePage_data = [{"firstname":"Bao Pham", "email":"baopham@udemy111.com", "gender":"Male"},{"firstname":"Tram Truong", "email":"Truong.Tram@whateverkjd.com", "gender":"Female"}]

        @staticmethod
        def getTestData(test_case_name):
                book = openpyxl.load_workbook("C:\\Users\\Main-Snow\\PycharmProjects\\PythonSelfFramework\\TestData\\PythonDemo.xlsx")
                sheet = book.active
                Dict = {}

                for i in range(1, sheet.max_row + 1):
                        if sheet.cell(row=i, column=1).value == test_case_name:

                                for j in range(2, sheet.max_column + 1):
                                        # Dict["lastname"] = "Bao"
                                        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

                return[Dict]
